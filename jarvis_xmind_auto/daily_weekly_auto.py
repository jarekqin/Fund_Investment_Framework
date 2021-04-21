import numpy as np
import pandas as pd
# import pandas_datareader.data as web
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib as mpl
import mpl_finance as finance_mpl
import yfinance as yf
import os

import excel2img

import time
import datetime

import requests
import warnings

# from omk.core.vendor.RQData import RQData
from omk.interface import AbstractJob
from omk.events import Event
from omk.utils.const import FolderName, TODAY, ProcessDocs, ProcessType, EVENT, TIME

from jarvis.utils import FOLDER, mkdir, copy_to_jarvisoutput
import matplotlib.ticker as ticker

from datetime import datetime, timedelta
from WindPy import w

from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from os import remove, path

mpl.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体：解决plot不能显示中文问题
mpl.rcParams['axes.unicode_minus'] = False
sns.set(font_scale=1.5, font='SimHei')
warnings.filterwarnings("ignore")

# based_file = FOLDER.Syn_save
based_file = 'E:\\'
my_fred_api_key = '7e41a8c76559d6eed6b02fbeccd136fd'


class MorningAfterMetting:
    @staticmethod
    def get_main_global_contracts(date=None, init=True):
        if init:
            w.start()
            # RQData.init()
        temp = w.wset("sectorconstituent",
                      "date=%s;sectorid=1000015511000000" % pd.to_datetime(date).strftime('%Y-%m-%d'))
        data = pd.DataFrame(np.transpose(temp.Data), columns=temp.Fields)
        # 过滤合约
        filter_contracts1 = {'SP500小型': 'CME', '纳指100小型': 'CME',
                             '黄金': 'CMX', '轻质原油': 'NYM'}
        filter_contracts2 = 'MGCM|QOM|SGCZ|SGUM|QMK'
        final_data = pd.DataFrame()
        for contract_name, contract_tail in filter_contracts1.items():
            final_data = pd.concat([final_data, data[data.sec_name.str.contains(contract_name) &
                                                     data.wind_code.str.contains(contract_tail)
                                                     ]],
                                   axis=0)
        final_data = final_data[final_data.wind_code.str.contains(filter_contracts2) == False]
        for index, code in zip(final_data.wind_code.index.to_list(),
                               final_data.wind_code.to_list()):
            if code.split('.')[0].rfind('E') == len(code.split('.')[0]) - 1:
                final_data.wind_code[index] = '.'.join([code.split('.')[0][:-1], code.split('.')[-1]])
            else:
                final_data.wind_code[index] = '.'.join([code.split('.')[0].replace('E', ''), code.split('.')[-1]])

        returned_dict = {x: y for x, y in zip(final_data.wind_code.to_list(), final_data.sec_name.to_list())}
        returned_dict.pop('QMM21.NYM')
        returned_dict.update({'GCM21.CMX': '黄金'})
        returned_dict.update({'CLM21.NYM': '轻质原油'})
        returned_dict.update({'^VIX': 'VIX'})
        returned_dict.update({'^TNX': '10年期美债收益率'})
        returned_dict.update({'ZT=F': '美债2年期货'})
        returned_dict.update({'ZN=F': '美债10年期货'})
        returned_dict.update({'TIP': '美国TIPS通涨保值债券ETF'})
        returned_dict.update({'^DJI': 'Dow指数'})
        returned_dict.update({'^IXIC': 'NASDAQ指数'})
        returned_dict.update({'^GSPC': 'SP500指数'})
        returned_dict.update({'^FTSE': '英国富时100'})
        returned_dict.update({'^GDAXI': 'GDAXI(德国DAX)'})
        returned_dict.update({'^RUT': '罗素2000指数'})
        returned_dict.update({'^NDX': 'NASDAQ100指数'})
        returned_dict.update({'^N225': '日经指数'})

        return returned_dict

    @staticmethod
    def get_past_one_month_trendency(wind_code_name_dict, date=None):
        if len(wind_code_name_dict) == 0:
            raise ValueError('wind code list is empty!')
        elif isinstance(wind_code_name_dict, dict):
            wind_code_list = ','.join(list(wind_code_name_dict.keys()))

        date = pd.to_datetime(date).strftime('%Y-%m-%d')
        start_date = (pd.to_datetime(date).date() - timedelta(30)).strftime('%Y-%m-%d')
        # 获取wind数据
        temp1 = w.wsd("%s" % wind_code_list, "close", "%s" % start_date, "%s" % date, "")
        trendency_df = pd.DataFrame(np.transpose(temp1.Data), columns=temp1.Codes, index=temp1.Times).fillna(0)
        # pct from previous:16:00 to today 16:00
        temp2 = w.wsd("%s" % wind_code_list, "pct_chg", "%s" % start_date, "%s" % date, "")
        pct_df = pd.DataFrame(np.transpose(temp2.Data), columns=temp2.Codes, index=temp2.Times).fillna(0)

        # 更换列名
        pct_df.columns = [wind_code_name_dict[key] for key in pct_df.columns]
        trendency_df.columns = [wind_code_name_dict[key] for key in trendency_df.columns]

        return trendency_df, pct_df

    @staticmethod
    def plot_candle(wind_code, contract_name, start_date, end_date,
                    use_wind=True, df=None, save_path=based_file):
        # get data
        start_date, end_date = pd.to_datetime(start_date), pd.to_datetime(end_date)
        if use_wind:
            temp = w.wsd("%s" % wind_code, "open,high,low,close", "%s" % start_date.strftime('%Y-%m-%d'),
                         "%s" % end_date.strftime('%Y-%m-%d'), "")
            data = pd.DataFrame(np.transpose(temp.Data), index=temp.Times, columns=temp.Fields).dropna()
        else:
            data = df[['Open', 'High', 'Low', 'Close']]
            data.columns = data.columns.str.upper()

        fig, ax = plt.subplots(figsize=(45, 20))
        finance_mpl.candlestick2_ohlc(ax, opens=data.OPEN, highs=data.HIGH, lows=data.LOW,
                                      closes=data.CLOSE, width=0.35, colorup='r', colordown='green')
        # finance_mpl.candlestick_ohlc(ax,data,width=0.35, colorup='r', colordown='green')
        ax.set_title(
            '报告日:%s,%s合约K线图' % (data.index[-1].strftime('%Y-%m-%d'), wind_code),
            fontsize=35
        )
        ax.set_xlabel('时间(亚洲/上海时区)', fontsize=35)
        ax.set_ylabel('价格点数', fontsize=35)
        plt.xticks(range(0, len(data.index), 8), [x.strftime('%Y-%m-%d') for x in
                                                  data.index[list(range(0, len(data.index),
                                                                        8))]],
                   rotation=90,
                   fontsize=35)
        ax.set_yticklabels([round(x, 2) for x in ax.get_yticks()], fontsize=35)
        fig.subplots_adjust(bottom=0.2)  #
        if save_path is None:
            plt.show()
        else:
            mkdir(os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d')))
            temp_path = os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'))
            save_path = os.path.join(temp_path, '%s.png' % (contract_name))
            plt.savefig(save_path, bbox_inches='tight')

    @staticmethod
    def plot_curve(data, picture_name,
                   ylabel_name, xlabel_name, title_value,
                   pct_value, pct_change,
                   save_path=based_file, plot_col=None):

        def format_date(x, pos=None):
            if x < 0 or x > len(data.index) - 1:
                return ''
            return data.index[int(x)]

        data.index = pd.to_datetime(data.index)
        # data.index=[x.strftime('%Y-%m-%d') for x in data.index]
        fig = plt.figure(figsize=(30, 15))
        ax = fig.add_subplot(111)
        x = range(len(data.index))
        if plot_col is None:
            ax.plot(x, data.values, label='Close')
        else:
            ax.plot(x, data[plot_col])
        plt.title('报告日:%s,%s分时图,绝对数是:%0.2f,涨跌是:%0.2f,涨跌幅是:%0.2f' % (
            data.index[-1].strftime('%Y-%m-%d'),
            picture_name,
            title_value, pct_value,
            pct_change) + '%')
        ax.set_ylabel(ylabel_name)
        ax.set_xlabel(xlabel_name)
        ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_date))
        ax.xaxis.set_major_locator(ticker.MultipleLocator(48))
        # plt.xticks(range(0, len(data.index), 8), [x.strftime('%Y-%m-%d %H:%M:%S') for x in
        #                                           data.index[list(range(0, len(data.index),
        #                                                                 8))]],
        #            rotation=45)
        fig.autofmt_xdate()
        if save_path is None:
            plt.show()
        else:
            mkdir(os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d')))
            temp_path = os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'))
            save_path = os.path.join(temp_path, '%s.png' % (picture_name))
            plt.savefig(save_path, bbox_inches='tight')

    @staticmethod
    def plot_curve2(data, picture_name,
                    ylabel_name, xlabel_name, title_value,
                    pct_value, pct_change,
                    save_path=based_file, plot_col=None):

        def format_date(x, pos=None):
            if x < 0 or x > len(data.index) - 1:
                return ''
            return data.index[int(x)]

        data.index = pd.to_datetime(data.index)
        # data.index=[x.strftime('%Y-%m-%d') for x in data.index]
        fig = plt.figure(figsize=(30, 15))
        ax = fig.add_subplot(111)
        x = range(len(data.index))
        if plot_col is None:
            ax.plot(x, data.values, label='Close')
        else:
            ax.plot(x, data[plot_col])
        plt.title('报告日:%s,%s日图,绝对数是:%0.2f,涨跌是:%0.2f,涨跌幅是:%0.2f' % (
            data.index[-1].strftime('%Y-%m-%d'),
            picture_name,
            title_value, pct_value,
            pct_change) + '%')
        ax.set_ylabel(ylabel_name)
        ax.set_xlabel(xlabel_name)
        ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_date))
        ax.xaxis.set_major_locator(ticker.MultipleLocator(48))
        # plt.xticks(range(0, len(data.index), 8), [x.strftime('%Y-%m-%d %H:%M:%S') for x in
        #                                           data.index[list(range(0, len(data.index),
        #                                                                 8))]],
        #            rotation=45)
        fig.autofmt_xdate()
        ax.set_xticks
        if save_path is None:
            plt.show()
        else:
            mkdir(os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d')))
            temp_path = os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'))
            save_path = os.path.join(temp_path, '%s.png' % (picture_name))
            plt.savefig(save_path, bbox_inches='tight')

    @staticmethod
    def plot_point_curve(data, picture_name,
                         ylabel_name, xlabel_name, title_value,
                         pct_value, pct_change,
                         save_path=based_file, plot_col=None):

        def format_date(x, pos=None):
            if x < 0 or x > len(data.index) - 1:
                return ''
            return data.index[int(x)]

        data.index = pd.to_datetime(data.index).strftime('%Y-%m-%d')
        # data.index=[x.strftime('%Y-%m-%d') for x in data.index]
        fig = plt.figure(figsize=(30, 15))
        ax = fig.add_subplot(111)
        x = range(len(data.index))
        if plot_col is None:
            ax.plot(x, data.values, 'ro-', label='Close')
        else:
            ax.plot(x, data[plot_col])
        plt.title('报告日:%s,%s分时图,绝对数是:%0.2f,涨跌是:%0.4f,涨跌幅是:%0.2f' % (
            data.index[-1],
            picture_name,
            title_value, pct_value,
            pct_change * 100) + '%')
        ax.set_ylabel(ylabel_name)
        ax.set_xlabel(xlabel_name)
        ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_date))
        ax.xaxis.set_major_locator(ticker.MultipleLocator(48))
        # plt.xticks(range(0, len(data.index), 8), [x.strftime('%Y-%m-%d %H:%M:%S') for x in
        #                                           data.index[list(range(0, len(data.index),
        #                                                                 8))]],
        #            rotation=45)
        fig.autofmt_xdate()
        if save_path is None:
            plt.show()
        else:
            mkdir(os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d')))
            temp_path = os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'))
            save_path = os.path.join(temp_path, '%s.png' % (picture_name))
            plt.savefig(save_path, bbox_inches='tight')

    @staticmethod
    def get_contract_data(wind_code, end_date, period='3d', interval='1m'):
        end_date = pd.to_datetime(end_date)
        if end_date.isoweekday() not in [1, 6, 7]:
            start_date = end_date - timedelta(1)
        elif end_date.isoweekday() == 1:
            start_date = end_date - timedelta(3)
        else:
            print('Weekend!')
        if isinstance(wind_code, dict) or isinstance(wind_code, list):
            data = {}
            daily_data = {}
            for code in wind_code:
                data_model = yf.Ticker(code)
                if code != '^VIX':
                    try:
                        temp_data = data_model.history(period=period, interval=interval,
                                                       start=start_date.strftime('%Y-%m-%d'),
                                                       end=(end_date + timedelta(2)).strftime('%Y-%m-%d'))
                        data[code] = temp_data
                        daily_data[code] = data_model.history(end=(end_date + timedelta(2)).strftime('%Y-%m-%d'),
                                                              start=pd.to_datetime(datetime(end_date.year - 1,
                                                                                            end_date.month,
                                                                                            end_date.day)).strftime(
                                                                  '%Y-%m-%d'))
                    except requests.exceptions.ConnectionError:
                        counter = 0
                        while True:
                            temp_data = pd.DataFrame()
                            try:
                                temp_data = data_model.history(period=period, interval=interval,
                                                               start=start_date.strftime('%Y-%m-%d'),
                                                               end=(end_date + timedelta(2)).strftime('%Y-%m-%d'))
                            except requests.exceptions.ConnectionError:
                                counter += 1
                            if counter >= 30:
                                requests.exceptions.ConnectionError('无法请求到雅虎金融的数据!')
                            if temp_data.shape[0] > 0:
                                break
                        data[code] = temp_data
                        daily_data[code] = data_model.history(end=(end_date + timedelta(2)).strftime('%Y-%m-%d'),
                                                              start=pd.to_datetime(datetime(end_date.year - 1,
                                                                                            end_date.month,
                                                                                            end_date.day)).strftime(
                                                                  '%Y-%m-%d'))
                else:
                    try:
                        temp_data = data_model.history(period=period, interval=interval,
                                                       start=start_date.strftime('%Y-%m-%d'),
                                                       end=(end_date + timedelta(2)).strftime('%Y-%m-%d'))
                        data[code] = temp_data
                        daily_data[code] = data_model.history(end=(end_date + timedelta(2)).strftime('%Y-%m-%d'),
                                                              start=pd.to_datetime(datetime(end_date.year - 1,
                                                                                            end_date.month,
                                                                                            end_date.day)).strftime(
                                                                  '%Y-%m-%d'))
                    except requests.exceptions.ConnectionError:
                        counter = 0
                        while True:
                            time.sleep(10)
                            temp_data = pd.DataFrame()
                            try:
                                temp_data = data_model.history(period=period, interval=interval,
                                                               start=start_date.strftime('%Y-%m-%d'),
                                                               end=(end_date + timedelta(2)).strftime('%Y-%m-%d'))
                            except requests.exceptions.ConnectionError:
                                counter += 1
                            if counter >= 30:
                                requests.exceptions.ConnectionError('无法请求到雅虎金融的数据!')
                            if temp_data.shape[0] > 0:
                                break
                        data[code] = temp_data
                        daily_data[code] = data_model.history(end=(end_date + timedelta(2)).strftime('%Y-%m-%d'),
                                                              start=pd.to_datetime(datetime(end_date.year - 1,
                                                                                            end_date.month,
                                                                                            end_date.day)).strftime(
                                                                  '%Y-%m-%d'))
        else:
            data_model = yf.Ticker(wind_code)
            data = {}
            daily_data = {}
            if wind_code != '^VIX':
                try:
                    temp_data = data_model.history(period=period, interval=interval,
                                                   start=start_date.strftime('%Y-%m-%d'),
                                                   end=(end_date + timedelta(2)).strftime('%Y-%m-%d'))
                except requests.exceptions.ConnectionError:
                    counter = 0
                    while True:
                        time.sleep(10)
                        temp_data = pd.DataFrame()
                        try:
                            temp_data = data_model.history(period=period, interval=interval,
                                                           start=start_date.strftime('%Y-%m-%d'),
                                                           end=(end_date + timedelta(2)).strftime('%Y-%m-%d'))
                        except requests.exceptions.ConnectionError:
                            counter += 1
                        if counter >= 30:
                            requests.exceptions.ConnectionError('无法请求到雅虎金融的数据!')
                        if temp_data.shape[0] > 0:
                            break
                data[wind_code] = temp_data
                daily_data[wind_code] = data_model.history(end=(end_date + timedelta(2)).strftime('%Y-%m-%d'),
                                                           start=pd.to_datetime(datetime(end_date.year - 1,
                                                                                         end_date.month,
                                                                                         end_date.day)).strftime(
                                                               '%Y-%m-%d'))
            else:
                try:
                    temp_data = data_model.history(period=period, interval=interval,
                                                   start=start_date.strftime('%Y-%m-%d'),
                                                   end=(end_date + timedelta(2)).strftime('%Y-%m-%d'))
                except requests.exceptions.ConnectionError:
                    counter = 0
                    while True:
                        time.sleep(10)
                        temp_data = pd.DataFrame()
                        try:
                            temp_data = data_model.history(period=period, interval=interval,
                                                           start=start_date.strftime('%Y-%m-%d'),
                                                           end=(end_date + timedelta(2)).strftime('%Y-%m-%d'))
                        except requests.exceptions.ConnectionError:
                            counter += 1
                        if counter >= 30:
                            raise requests.exceptions.ConnectionError('无法请求到雅虎金融的数据!')
                        if temp_data.shape[0] > 0:
                            break
                data[wind_code] = temp_data
                daily_data[wind_code] = data_model.history(end=(end_date + timedelta(2)).strftime('%Y-%m-%d'),
                                                           start=pd.to_datetime(datetime(end_date.year - 1,
                                                                                         end_date.month,
                                                                                         end_date.day)).strftime(
                                                               '%Y-%m-%d'))
        # we will cut date range from T-1 at 8a.m to T at 8a.m from data according to time zone different
        if len(data.keys()) > 0:
            # TODO modify pct_change,pct_value
            returned_dict = {}
            for key in data:
                print(key)
                data[key].index = [x.tz_convert('Asia/Shanghai') for x in data[key].index]
                if key in ['TIP', '^DJI', '^IXIC', '^GSPC', '^NDX', '^RUT', '^GDAXI', '^FTSE']:
                    returned_dict[key + '_晨会'] = {
                        '涨跌': data[key].loc[end_date.strftime('%Y-%m-%d')].Close[-1] - \
                              data[key].loc[start_date.strftime('%Y-%m-%d')].Close[-1],
                        '涨跌幅': round((data[key].loc[end_date.strftime('%Y-%m-%d')].Close[-1] - \
                                      data[key].loc[start_date.strftime('%Y-%m-%d')].Close[-1]) / \
                                     data[key].loc[start_date.strftime('%Y-%m-%d')].Close[-1] * 100, 2),
                        '收盘绝对数': data[key].loc[end_date.strftime('%Y-%m-%d')].Close[-1],
                        '分时图数据': data[key]
                    }
                elif key in ['ZT=F', 'ZN=F', 'ESM21.CME', 'NQM21.CME', 'CLM21.NYM']:
                    if datetime.now().hour > 23:
                        temp_data = data[key].loc[(end_date + timedelta(1)).strftime('%Y-%m-%d 08:00:00'):(
                                end_date + timedelta(1)).strftime('%Y-%m-%d 16:00:00')]
                        returned_dict[key + '_复盘'] = {
                            '涨跌': data[key].Close[-1] - data[key].loc[
                                (end_date + timedelta(1)).strftime('%Y-%m-%d 08:00:00')].Close,
                            '涨跌幅': round(
                                (data[key].Close[-1] - data[key].loc[
                                    (end_date + timedelta(1)).strftime('%Y-%m-%d 08:00:00')].Close) / \
                                data[key].loc[(end_date + timedelta(1)).strftime('%Y-%m-%d 08:00:00')].Close * 100, 2),
                            '收盘绝对数': data[key].Close[-1],
                            '分时图数据': temp_data
                        }
                    else:
                        temp_data = data[key].loc[
                                    end_date.strftime('%Y-%m-%d 15:00:00'):(end_date + timedelta(1)).strftime(
                                        '%Y-%m-%d 08:00:00')]
                        returned_dict[key + '_晨会'] = {
                            '涨跌': data[key].Close[-1] - data[key].loc[end_date.strftime('%Y-%m-%d 15:00:00')].Close,
                            '涨跌幅': round(
                                (data[key].Close[-1] - data[key].loc[end_date.strftime('%Y-%m-%d 15:00:00')].Close) / \
                                data[key].loc[end_date.strftime('%Y-%m-%d 15:00:00')].Close * 100, 2),
                            '收盘绝对数': data[key].Close[-1],
                            '分时图数据': temp_data
                        }
                elif key in ['GCM21.CMX', 'CLK21.NYM', '^VIX']:
                    if key != '^VIX':
                        if datetime.now().hour > 23:
                            temp_data = data[key].loc[(end_date + timedelta(1)).strftime('%Y-%m-%d 08:00:00'):(
                                    end_date + timedelta(1)).strftime('%Y-%m-%d 16:00:00')]
                            returned_dict[key + '_复盘'] = {
                                '涨跌': data[key].Close[-1] - data[key].loc[
                                    (end_date + timedelta(1)).strftime('%Y-%m-%d 08:00:00')].Close,
                                '涨跌幅': round(
                                    (data[key].Close[-1] - data[key].loc[
                                        (end_date + timedelta(1)).strftime('%Y-%m-%d 08:00:00')].Close) / \
                                    data[key].loc[(end_date + timedelta(1)).strftime('%Y-%m-%d 08:00:00')].Close * 100,
                                    2),
                                '收盘绝对数': data[key].Close[-1],
                                '分时图数据': temp_data
                            }
                        else:
                            temp_data = data[key].loc[
                                        end_date.strftime('%Y-%m-%d 15:00:00'):(end_date + timedelta(1)).strftime(
                                            '%Y-%m-%d 08:00:00')]
                            returned_dict[key + '_晨会'] = {
                                '涨跌': data[key].Close[-1] - data[key].loc[end_date.strftime('%Y-%m-%d 15:00:00')].Close,
                                '涨跌幅': round(
                                    (data[key].Close[-1] - data[key].loc[
                                        end_date.strftime('%Y-%m-%d 15:00:00')].Close) / \
                                    data[key].loc[end_date.strftime('%Y-%m-%d 15:00:00')].Close * 100, 2),
                                '收盘绝对数': data[key].Close[-1],
                                '分时图数据': temp_data
                            }
                    else:
                        if datetime.now().hour > 23:
                            returned_dict[key + '_复盘'] = {
                                '涨跌': data[key].Close[-1] - data[key].Close[0],
                                '涨跌幅': round((data[key].Close[-1] - data[key].Close[0]) / data[key].Close[0] * 100, 2),
                                '收盘绝对数': data[key].Close[-1],
                                '分时图数据': data[key]
                            }
                        else:
                            returned_dict[key + '_晨会'] = {
                                '涨跌': data[key].Close[-1] - data[key].Close[0],
                                '涨跌幅': round((data[key].Close[-1] - data[key].Close[0]) / data[key].Close[0] * 100, 2),
                                '收盘绝对数': data[key].Close[-1],
                                '分时图数据': data[key]
                            }
                elif key == '^N225':
                    if datetime.now().hour > 23:
                        returned_dict[key + '_复盘'] = {
                            '涨跌': data[key].Close[-1] - data[key].Close[0],
                            '涨跌幅': round((data[key].Close[-1] - data[key].Close[0]) / data[key].Close[0] * 100, 2),
                            '收盘绝对数': data[key].Close[-1],
                            '分时图数据': data[key]
                        }
                    else:
                        returned_dict[key + '_晨会'] = {
                            '涨跌': data[key].Close[-1] - data[key].Close[0],
                            '涨跌幅': round((data[key].Close[-1] - data[key].Close[0]) / data[key].Close[0] * 100, 2),
                            '收盘绝对数': data[key].Close[-1],
                            '分时图数据': data[key]
                        }
                elif key == '^TNX':
                    returned_dict[key + '_晨会'] = {
                        '涨跌': data[key].Close[-1] - data[key].Close[0],
                        '涨跌幅': round((data[key].Close[-1] - data[key].Close[0]) / data[key].Close[0] * 100, 2),
                        '收盘绝对数': data[key].Close[-1],
                        '分时图数据': data[key]
                    }
                else:
                    print('%s不需要提取数据' % code)
        else:
            raise RuntimeError('没有抓到任何数据!')

        return returned_dict, daily_data

    @staticmethod
    def get_wind_k_plot_code(end_date, code_list=None):
        end_date = pd.to_datetime(end_date)
        if code_list is None:
            code_list = {'USDX.FX': '美元指数', 'USDCNH.FX': '美元兑换离岸人名币', 'SX5E.GI': '斯托克50欧元'}

        temp_data = w.wset("sectorconstituent", "date=%s;sectorid=1000015510000000" % end_date.strftime('%Y-%m-%d'))
        filter_data = pd.DataFrame(np.transpose(temp_data.Data), columns=temp_data.Fields)
        filter_list = ['T%s' % str(end_date.year)[-2:], 'TS%s' % str(end_date.year)[-2:]]
        filter_contract = filter_data[
            filter_data.sec_name.str.contains('|'.join([x for x in filter_list]))][['wind_code', 'sec_name']]
        for code, name in zip(filter_contract.wind_code, filter_contract.sec_name):
            code_list[code] = name

        # taking data from wind
        # final_data = {}
        # for contract in code_list:
        #     temp_data=w.wsd("%s" % contract, "open,high,low,close,pct_chg,chg", "%s" % start_date.strftime('%Y-%m-%d'),
        #           "%s" % end_date.strftime('%Y-%m-%d'), "")
        #     data=pd.DataFrame(np.transpose(temp_data.Data),columns=temp_data.Fields,index=temp_data.Times)
        #     final_data[contract]=data

        return code_list

    @staticmethod
    def get_fixed_data_from_wind(wind_code, data_col, start_date, end_date):
        start_date, end_date = pd.to_datetime(start_date), pd.to_datetime(end_date)
        if isinstance(wind_code, list):
            wind_code = ','.join([x for x in wind_code])
        else:
            wind_code = wind_code
        if isinstance(data_col, list):
            data_col = ','.join([x for x in data_col])
        else:
            data_col = data_col
        temp_data = w.wsd("%s" % wind_code, data_col,
                          "%s" % start_date.strftime('%Y-%m-%d'),
                          "%s" % end_date.strftime('%Y-%m-%d'), "")
        data = pd.DataFrame(np.transpose(temp_data.Data),
                            index=temp_data.Times, columns=temp_data.Fields)
        return data

    @staticmethod
    def get_fred_data(end_date, init=True):
        end_date = pd.to_datetime(end_date)
        if init:
            import fred
            fred.key(my_fred_api_key)
        original_data = pd.DataFrame(fred.observations('T10YIE')['observations']).set_index('date')
        original_data.index = pd.to_datetime(original_data.index)
        start_date = datetime(end_date.year - 1, end_date.month, end_date.day)
        return_data = original_data.loc[start_date:end_date]
        return_data = return_data.replace('.', np.nan).dropna()
        return return_data.value.apply(float)

    @staticmethod
    def get_currency_output(end_date=None):
        if end_date is None:
            end_date = datetime.today()
        else:
            end_date = pd.to_datetime(end_date)
        start_date = end_date - timedelta(60)
        # code is fixed
        temp_data = w.edb("M0041372,M0062600,M6217188,M0329542,M5528819",
                          "%s" % start_date.strftime('%Y-%m-%d'), "%s" % end_date.strftime('%Y-%m-%d'),
                          "Fill=Previous")
        data = pd.DataFrame(np.transpose(temp_data.Data), index=temp_data.Times,
                            columns=temp_data.Codes)
        data.columns = ['逆回购7天', '逆回购到期', 'TMLF投放',
                        'MLF投放', 'MLF收回']

        return data

    @staticmethod
    def get_bond_ytm_from_wind(end_date=None):
        if end_date is None:
            end_date = datetime.today()
        else:
            end_date = pd.to_datetime(end_date)
        start_date = datetime(end_date.year - 1, end_date.month, end_date.day)
        temp_data = w.edb("M1001654,M1001647", "%s" % start_date.strftime('%Y-%m-%d'),
                          "%s" % end_date.strftime('%Y-%m-%d'), "Fill=Previous")
        data = pd.DataFrame(np.transpose(temp_data.Data), index=temp_data.Times,
                            columns=temp_data.Codes)
        return data

    @staticmethod
    def get_wind_tick_data(code, start_date, end_date):
        start_date, end_date = pd.to_datetime(start_date), pd.to_datetime(end_date)
        temp_data = w.wsi("%s" % code, "close,open", "%s 09:00:00" % start_date.strftime('%Y-%m-%d'),
                          "%s 15:59:00" % end_date.strftime('%Y-%m-%d'), "")
        data = pd.DataFrame(np.transpose(temp_data.Data), index=temp_data.Times, columns=temp_data.Fields)
        return data.fillna(method='ffill')

    @staticmethod
    def get_yahoo_daily_data(code, start_date, end_date):
        start_date, end_date = pd.to_datetime(start_date), pd.to_datetime(end_date)
        model = yf.Ticker(code)
        his_data = model.history(start=start_date, end=end_date)
        his_data.index = [x.tz_localize('Asia/Shanghai') for x in his_data.index]
        return his_data

    @staticmethod
    def get_north_south_data_from_wind(start_date, end_date):
        start_date, end_date = pd.to_datetime(start_date), pd.to_datetime(end_date)
        # top line
        temp_data = w.wset("shhkactivitystock", "startdate=%s;enddate=%s;direction=south" %
                           (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        south = pd.DataFrame((np.transpose(temp_data.Data)), columns=temp_data.Fields)
        # north
        temp_data = w.wset("shhkactivitystock", "startdate=%s;enddate=%s;direction=north" %
                           (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        north = pd.DataFrame((np.transpose(temp_data.Data)), columns=temp_data.Fields)

        cols = ['代码', '证券名', '日期', '类型', '排名', '买卖总额', '买入', '卖出', '成交净价', '收盘价',
                '货币', '涨跌', '涨跌幅', '换手率', '市盈率', '市净率', '交易所行业', '行业']
        # to excel
        south.columns = cols
        north.columns = cols

        south = south.set_index('日期')
        north = north.set_index('日期')

        south.index = [x.strftime('%Y%m%d') for x in south.index]
        north.index = [x.strftime('%Y%m%d') for x in north.index]

        south.index.name = '日期'
        north.index.name = '日期'

        file_path = os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'),
                                 'south_north_data.xlsx')
        xlsx = XlsxSaver(north.reset_index(), file_path, 'north')
        for col in cols:
            if col == '交易所行业':
                xlsx.set_width(col, 25)
            else:
                xlsx.set_width(col, 15)
        xlsx.save()

        xlsx = XlsxSaver(south.reset_index(), file_path, 'south')
        for col in cols:
            if col == '交易所行业':
                xlsx.set_width(col, 25)
            else:
                xlsx.set_width(col, 15)
        xlsx.save()

        excel2img.export_img(os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'),
                                          'south_north_data.xlsx'),
                             os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'),
                                          'south.png'),
                             '', 'south!A1:R41')

        excel2img.export_img(os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'),
                                          'south_north_data.xlsx'),
                             os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'),
                                          'north.png'),
                             '', 'north!A1:R41')

    @staticmethod
    def main(end_date=None):
        if end_date is None:
            wind_code_name_dict = MorningAfterMetting.get_main_global_contracts(end_date=datetime.today())
            returned_df, daily_data = MorningAfterMetting.get_contract_data(wind_code_name_dict, datetime.now())
        else:
            wind_code_name_dict = MorningAfterMetting.get_main_global_contracts(date=pd.to_datetime(end_date).date())
            returned_df, daily_data = MorningAfterMetting.get_contract_data(wind_code_name_dict,
                                                                            pd.to_datetime(end_date).to_pydatetime())

        # plot candle plots1
        code_list = MorningAfterMetting.get_wind_k_plot_code(end_date)
        for code in code_list:
            MorningAfterMetting.plot_candle(wind_code=code, contract_name=code_list[code],
                                            start_date=datetime(end_date.year - 1, end_date.month,
                                                                end_date.day).strftime(
                                                '%Y-%m-%d'),
                                            end_date=end_date.strftime('%Y-%m-%d'))

        # chinese bond minuts plot
        chinese_bond_list = [x for x in code_list if 'T' in x]
        for bond_code in chinese_bond_list:
            data = MorningAfterMetting.get_wind_tick_data(bond_code, (end_date - timedelta(1)), end_date)
            MorningAfterMetting.plot_curve(data, bond_code, '收盘价', '时间(亚洲/上海时区)',
                                           data.close[-1],
                                           data.close[-1] -
                                           data.loc[(end_date - timedelta(1)).strftime('%Y-%m-%d')].close[0],
                                           (data.close[-1] - data.loc[(end_date - timedelta(1)).strftime('%Y-%m-%d')]
                                            .close[0]) / data.close[0],
                                           plot_col='close')
        # data, picture_name,
        # ylabel_name, xlabel_name, title_value,
        # pct_value, pct_change,
        # save_path = based_file, plot_col = None
        # plot candle plots2
        for code in daily_data:
            temp_data = daily_data[code]
            MorningAfterMetting.plot_candle(wind_code_name_dict[code.split('_')[0]],
                                            wind_code_name_dict[code.split('_')[0]],
                                            end_date.strftime('%Y-%m-%d'), None, False, temp_data, based_file)

        # plot minute plots
        # taking morning code
        # if datetime.now().hour < 16 and datetime.now().hour >= 8:
        if datetime.now().hour < 23:
            morning_list = ['^VIX_晨会', '^TNX_晨会', 'TIP_晨会', '^DJI_晨会', '^IXIC_晨会', '^GSPC_晨会', '^FTSE_晨会',
                            '^GDAXI_晨会', 'GCM21.CMX_晨会', 'CLM21.NYM_晨会', '^N225_晨会']
            for code in morning_list:
                MorningAfterMetting.plot_curve(
                    returned_df[code]['分时图数据']['Close'],
                    code.replace('_晨会', ''), '收盘价', '时间(亚洲/上海时区)',
                    returned_df[code]['收盘绝对数'],
                    returned_df[code]['涨跌'],
                    returned_df[code]['涨跌幅']
                )

            #
        elif datetime.now().hour > 23:
            afternoon_list = ['^N225_复盘', 'ZT=F_复盘', 'ZN=F_复盘', 'GCM21.CMX_复盘', 'CLM21.NYM_复盘',
                              'ESM21.CME_复盘', 'NQM21.CME_复盘']
            for code in afternoon_list:
                MorningAfterMetting.plot_curve(
                    returned_df[code]['分时图数据']['Close'],
                    code.replace('_复盘', ''), '收盘价', '时间(亚洲/上海时区)',
                    returned_df[code]['收盘绝对数'],
                    returned_df[code]['涨跌'],
                    returned_df[code]['涨跌幅']
                )
        else:
            print('当前时间段非8~16点！')

        # solving factor from wind
        wind_code = ['DR007.IB']
        for code in wind_code:
            data = MorningAfterMetting.get_fixed_data_from_wind(wind_code=code, data_col=['open', 'close'],
                                                                start_date=datetime(end_date.year - 1, end_date.month,
                                                                                    end_date.day),
                                                                end_date=end_date
                                                                )
            title_value = data.CLOSE[-1]
            pct_value = data.CLOSE.diff()[-1]
            pct_change = data.CLOSE.pct_change()[-1]
            MorningAfterMetting.plot_curve2(data, 'DR007',
                                            '银行质押7天利率', '时间(亚洲/上海时区)',
                                            title_value, pct_value,
                                            pct_change,
                                            save_path=based_file, plot_col='CLOSE')

        # fred infaltion rate 10 years
        fred_data = MorningAfterMetting.get_fred_data(end_date)
        # title_value = fred_data.iloc[-1] - fred_data.iloc[0]

        MorningAfterMetting.plot_point_curve(fred_data, 'T10YIE',
                                             '美国十年期通涨盈亏平衡率', '时间',
                                             fred_data.iloc[-1], fred_data.iloc[-1] - fred_data.iloc[-2],
                                             (fred_data.iloc[-1] - fred_data.iloc[-2]) / fred_data.iloc[-2],
                                             save_path=based_file)

        # hedge_data = MorningAfterMetting.get_currency_output(end_date)
        # hedge_data.index = [x.strftime('%Y%m%d') for x in hedge_data.index]
        # hedge_data.to_excel(os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'),
        #                                  'central_bank_hedge_data.xlsx'), encoding='utf8')
        # row_num = hedge_data.shape[0]
        # excel2img.export_img(os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'),
        #                                   'central_bank_hedge_data.xlsx'),
        #                      os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'), 'OMO.png'),
        #                      '', 'Sheet1!A1:F%d' % (row_num + 1))

        # plot line for bond ytm
        bond_data = MorningAfterMetting.get_bond_ytm_from_wind(end_date)
        bond_data_diff = bond_data.iloc[:, 0] - bond_data.iloc[:, -1]
        bond_data_diff.name = '10年-2年收益率差'
        MorningAfterMetting.plot_point_curve(
            bond_data_diff, '中国10年国债-中国2年国债收益率差值',
            '10年-2年收益率差',
            '时间(亚洲/上海时区)', bond_data_diff.iloc[-1],
            bond_data_diff.iloc[-1] - bond_data_diff.iloc[-2],
            bond_data_diff.pct_change().iloc[-1]
        )

        # nasdaq100/rusell2000
        nasdaq100_data = MorningAfterMetting.get_yahoo_daily_data('^NDX', datetime(end_date.year - 1,
                                                                                   end_date.month,
                                                                                   end_date.day),
                                                                  end_date + timedelta(1))
        rusell2000_data = MorningAfterMetting.get_yahoo_daily_data('^RUT', datetime(end_date.year - 1,
                                                                                    end_date.month,
                                                                                    end_date.day),
                                                                   end_date + timedelta(1))
        divided_data = nasdaq100_data.Close / rusell2000_data.Close

        MorningAfterMetting.plot_point_curve(divided_data, 'NASDAQ100VSRUSSELL2000', '收盘价', '时间(亚洲/上海时区)',
                                             divided_data.iloc[-1],
                                             divided_data.iloc[-1] - divided_data.iloc[-2],
                                             (divided_data.iloc[-1] - divided_data.iloc[-2]) / divided_data.iloc[-2])

        MorningAfterMetting.get_north_south_data_from_wind(datetime(end_date.year - 1,
                                                                    end_date.month,
                                                                    end_date.day),
                                                           end_date)

        # 美国10年国债收益率折线图
        code = '^TNX'
        model = yf.Ticker(code)
        his_data = model.history(start=datetime(end_date.year - 1, end_date.month, end_date.day).strftime('%Y-%m-%d'),
                                 end=end_date.strftime('%Y-%m-%d'))
        MorningAfterMetting.plot_point_curve(his_data, '10_Year_YTM_curve',
                                             '利率收盘价', '时间', his_data.Close[-1],
                                             his_data.Close[-1] - his_data.Close[-2],
                                             (his_data.Close[-1] - his_data.Close[-2]) / his_data.Close[-2],
                                             plot_col='Close')

        # 10 YEAR USA NOTE - INFATION RATE
        fred_data2 = fred_data.loc[his_data.index]
        result = his_data.Close - fred_data2
        result.name = '10 Year Bond YTM - 10 Year Inflation Rate'
        MorningAfterMetting.plot_point_curve(result, '10 Year Bond YTM - 10 Year Inflation Rate',
                                             '利率差值', '时间',
                                             result.iloc[-1],
                                             result.iloc[-1] - result.iloc[-2],
                                             np.nan)


class WeekendMeeting:
    @staticmethod
    def index_pct_change(code, start_date=None, end_date=datetime.today().date()):
        if isinstance(code, list):
            code = ','.join([x for x in code])
        else:
            code = code
        if start_date is None:
            start_date = datetime(2021, 1, 1)
        start_date, end_date = pd.to_datetime(start_date), pd.to_datetime(end_date)
        temp_data = w.wsd("%s" % code, "pct_chg", "%s" % start_date.strftime('%Y-%m-%d'),
                          "%s" % end_date.strftime('%Y-%m-%d'), "")
        data = pd.DataFrame(np.transpose(temp_data.Data), index=temp_data.Times, columns=temp_data.Codes)
        return data

    @staticmethod
    def plot_point_curve(data, picture_name, ylabel_name, xlabel_name,
                         plot_col=None, save_path=based_file):

        def format_date(x, pos=None):
            if x < 0 or x > len(data.index) - 1:
                return ''
            return data.index[int(x)]

        data.index = pd.to_datetime(data.index).strftime('%Y-%m-%d')
        # data.index=[x.strftime('%Y-%m-%d') for x in data.index]
        if not isinstance(data, pd.DataFrame):
            fig = plt.figure(figsize=(30, 15))
            ax = fig.add_subplot(111)
            x = range(len(data.index))
            if plot_col is None:
                ax.plot(x, data.values, 'o-', label='Close')
            else:
                ax.plot(x, data[plot_col])
            plt.title('报告日:%s,%s' % (data.index[-1], picture_name))
            ax.set_ylabel(ylabel_name)
            ax.set_xlabel(xlabel_name)
            ax.grid(True)
            ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_date))
            ax.xaxis.set_major_locator(ticker.MultipleLocator(48))
            # plt.xticks(range(0, len(data.index), 8), [x.strftime('%Y-%m-%d %H:%M:%S') for x in
            #                                           data.index[list(range(0, len(data.index),
            #                                                                 8))]],
            #            rotation=45)
            fig.autofmt_xdate()
        else:
            ax = data[plot_col].plot(figsize=(30, 15), secondary_y=True,
                                     fontsize=20, x_compat=True,
                                     mark_right=False,
                                     legend=False, grid=True
                                     )
            plt.ylabel(ylabel_name, fontsize=20)
            plt.xlabel(xlabel_name, fontsize=20)
            plt.title('报告日:%s,%s' % (data.index[-1], picture_name))
            plt.grid(True)
            plt.legend(loc='best')
        if save_path is None:
            plt.show()
        else:
            mkdir(os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d')))
            temp_path = os.path.join(based_file, 'morning_temp', datetime.today().strftime('%Y-%m-%d'))
            save_path = os.path.join(temp_path, '%s.png' % (picture_name))
            plt.savefig(save_path, bbox_inches='tight')

    @staticmethod
    def copy_funds_plot(key_words, date, read_path, save_file_name='morning_temp'):
        date = pd.to_datetime(date)
        whole_files = []
        for root, dir, files in os.walk(read_path):
            for file in files:
                whole_files.append(os.path.join(root, file))

        if isinstance(key_words, list):
            key_words = [date.strftime('%Y-%m-%d') + x if '_Fund_issued_analyst' in x else x for x in key_words]
        else:
            if '_Fund_issued_analyst' in key_words:
                key_words = date.strftime('%Y-%m-%d') + '_Fund_issued_analyst'
            else:
                key_words = key_words

        try:
            copy_to_jarvisoutput(key_words, save_file_name=save_file_name, read_based_folder=read_path)
        except Exception as e:
            raise e

    @staticmethod
    def main(start_date=None, end_date=None, weekend_day=None,
             key_words=['Fund_issued_analyst.png', '限售解禁VS高层减持规模.png'],
             save_file_name='morning_temp'):
        if pd.to_datetime(end_date).isoweekday() not in [6, 7]:
            raise ValueError('end date is not weekend!')
        else:
            if datetime.today().isoweekday() == 6:
                end_date = pd.to_datetime(end_date) - timedelta(1)
            elif datetime.today().isoweekday() == 7:
                end_date = pd.to_datetime(end_date) - timedelta(2)
            else:
                end_date = pd.to_datetime(end_date)
        if start_date is None:
            start_date = pd.to_datetime(datetime(2021, 1, 1))
        # main indeies around world
        index_code = ['000001.SH', '399006.SZ', 'HSI.HI', 'N225.GI', 'SPX.GI', 'IXIC.GI', 'SX5E.GI']
        index_pct_change_data = WeekendMeeting.index_pct_change(index_code, start_date, end_date)
        WeekendMeeting.plot_point_curve(((index_pct_change_data / 100).cumsum() + 1).fillna(method='ffill'),
                                        '全球重要指数今年以来表现', '指数累计收益率(今年以来),起点为1',
                                        '时间(国内时区)', index_pct_change_data.columns.to_list())

        WeekendMeeting.copy_funds_plot(key_words, weekend_day, os.path.join(FOLDER.Syn_read, 'Fund_Issue_Reports'),
                                       save_file_name=save_file_name)


class XlsxSaver:
    """
    一个将DataFrame转换成格式化excel的工具
    """

    def __init__(self, df_in, filename='a.xlsx', sheet_name='Sheet1'):
        """
        df_in : 从一个DataFrame对象获取表格内容
        filename : 文件名
        sheet_name : 表名
        """
        self.filename = filename  # 保存的xlsx文件的名字
        self.user_def = []  # 储存由用户自定义的列的列名，这些列不再参与自动计算列宽
        if path.exists(filename):
            # 如果文件存在，就直接打开，添加Sheet
            self.wb = load_workbook(filename)
            self.sheet = self.wb.create_sheet(sheet_name)
        else:
            # 如果文件不存在，就创建表格
            self.wb = Workbook()
            self.sheet = self.wb.active
            self.sheet.title = sheet_name
        # 将df的内容复制给sheet
        self.df = df_in.copy()
        self.sheet.append(list(self.df.columns))
        for row in range(0, len(list(self.df.index))):
            for col in range(0, len(list(self.df.columns))):
                self.sheet.cell(row + 2, col + 1).value = self.df.iloc[row, col]  # 注意：sheet行列从1开始计数

    def remove_file(self):
        remove(self.filename)

    def set_sheet_name(self, sheet_name):
        self.sheet.title = sheet_name

    def set_filename(self, filename):
        self.filename = filename

    def get_maxlength(self, series_in, col):
        """
        获取一个类型为object的Series中的最大占位长度，用于确定导出的xlsx文件的列宽
        col : 表头，也参与比较，解决有时候表头过长的问题
        """
        series = series_in.fillna('-')  # 填充空值，防止出现nan
        str_list = list(series)
        len_list = []
        for elem in str_list + [col]:
            elem_split = list(elem)
            length = 0
            for c in elem_split:
                if ord(c) <= 256:
                    length += 1
                else:
                    length += 2
            len_list.append(length)
        return max(len_list)

    def __auto_width(self):
        cols_list = list(self.df.columns)  # 获取列名
        for i in range(0, len(cols_list)):
            col = cols_list[i]
            if col in self.user_def:
                continue
            self.sheet.cell(1, i + 1).font = Font(bold=True)  # 加粗表头
            letter = chr(i + 65)  # 由ASCII值获得对应的列字母
            max_len = self.get_maxlength(self.df[col].astype(str), col)
            if max_len <= 12:
                self.sheet.column_dimensions[letter].width = 12
            elif max_len <= 50:
                self.sheet.column_dimensions[letter].width = max_len + 2
            else:
                self.sheet.column_dimensions[letter].width = 50
                for cell in self.sheet[letter]:
                    cell.alignment = Alignment(wrap_text=True)

    def set_width(self, col_name, width):
        # 提供调整列宽的接口
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        self.sheet.column_dimensions[letter].width = width
        self.user_def.append(col_name)

    def set_color(self, col_name, color, rule):
        # 提供设置颜色的接口，rule:规则函数
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        for cell in self.sheet[letter]:
            if rule(cell.value):
                cell.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)

    def set_center_alignment(self, col_name):
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        for cell in self.sheet[letter]:
            cell.alignment = Alignment(wrap_text=True, horizontal='center')

    def save(self):
        # 自动调整列宽，并保存
        self.__auto_width()
        self.wb.save(self.filename)

    def set_merge(self, col_name):
        self.user_def.append(col_name)  # 设置为自定义列
        # 设置一列合并单元格
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        i = 1
        while True:
            if i >= self.sheet.max_row:
                # 结束条件：单元格到底
                break
            cell = self.sheet[letter + str(i)]
            j = i + 1  # 第一步指向下一个单元格
            while True:
                # 这个过程对j进行试探，最终j指向的单元格是与i连续相同的最后一个
                cell_next = self.sheet[letter + str(j)]
                if cell_next.value != cell.value:
                    j -= 1
                    break
                else:
                    j += 1
                if j > self.sheet.max_row:
                    j -= 1
                    break
            if j - i >= 1 and cell.value != '' and cell.value:
                # 如果有连续两格以上的单元格内容相同，进行融合
                msg = '%s%d:%s%d' % (letter, i, letter, j)
                self.sheet.merge_cells(msg)
            # 控制一下格式
            self.sheet[letter + str(i)].alignment = Alignment(horizontal='center',
                                                              vertical='top',
                                                              wrap_text=True)
            i = j + 1  # 继续指向下个单元格


if __name__ == '__main__':
    w.start()
    if datetime.today().isoweekday() not in [1, 7, 6]:
        MorningAfterMetting.main(datetime.today() - timedelta(1))
    elif datetime.today().isoweekday() == 1:
        MorningAfterMetting.main(datetime.today() - timedelta(3))
    else:
        WeekendMeeting.main(start_date=None, end_date=datetime(2021, 4, 24), weekend_day='2021-04-16')
